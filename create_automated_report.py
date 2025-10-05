#!/usr/bin/env python3
"""
🤖 АВТОМАТИЧЕСКИЙ ОТЧЁТ С ГРАФИКАМИ
Создаёт Excel файл с встроенными PNG графиками
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pathlib import Path
import re
from datetime import datetime
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill
import warnings
warnings.filterwarnings('ignore')

# Настройка matplotlib для русских шрифтов
plt.rcParams['font.family'] = ['DejaVu Sans', 'Arial Unicode MS', 'sans-serif']
plt.rcParams['axes.unicode_minus'] = False

def create_summary_chart(report_dir):
    """Создаёт сводный график для отчёта"""
    print("📊 Создаём сводный график для отчёта...")
    
    # Загружаем данные
    df_26sep, df_5oct = load_data_by_date()
    df_26sep = clean_salary_data(df_26sep, "26 сентября")
    df_5oct = clean_salary_data(df_5oct, "5 октября")
    
    # Вычисляем статистики
    stats_26sep = get_stats(df_26sep, "26 сентября")
    stats_5oct = get_stats(df_5oct, "5 октября")
    
    # Создаём сводный график
    fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(16, 12))
    
    # График 1: Количество вакансий
    dates = ['26 сентября\n(24-26.09)', '5 октября\n(3-5.10)']
    values = [stats_26sep['with_salary'], stats_5oct['with_salary']]
    bars1 = ax1.bar(dates, values, color=['lightblue', 'lightcoral'])
    ax1.set_title('📊 Количество вакансий с зарплатой', fontsize=14, fontweight='bold')
    ax1.set_ylabel('Количество вакансий')
    
    for i, bar in enumerate(bars1):
        height = bar.get_height()
        ax1.text(bar.get_x() + bar.get_width()/2., height + 5,
                f'{int(height)}', ha='center', va='bottom', fontweight='bold')
    
    # Темп роста
    delta = ((stats_5oct['with_salary'] - stats_26sep['with_salary']) / stats_26sep['with_salary']) * 100
    arrow = '↑' if delta > 0 else '↓' if delta < 0 else '→'
    color = 'green' if delta > 0 else 'red' if delta < 0 else 'gray'
    
    mid_x = (bars1[0].get_x() + bars1[0].get_width() + bars1[1].get_x()) / 2
    max_y = max(values) + 15
    
    ax1.text(mid_x, max_y, f'{arrow} {delta:+.1f}%', 
            ha='center', va='bottom', fontweight='bold', fontsize=12, color=color)
    
    ax1.grid(True, alpha=0.3)
    
    # График 2: Средние зарплаты
    values2 = [stats_26sep['mean_salary'], stats_5oct['mean_salary']]
    bars2 = ax2.bar(dates, values2, color=['lightgreen', 'orange'])
    ax2.set_title('💰 Средние зарплаты', fontsize=14, fontweight='bold')
    ax2.set_ylabel('Зарплата, ₽')
    
    for i, bar in enumerate(bars2):
        height = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2., height + 5000,
                f'{height:,.0f} ₽', ha='center', va='bottom', fontweight='bold')
    
    # Темп роста зарплат
    delta2 = ((stats_5oct['mean_salary'] - stats_26sep['mean_salary']) / stats_26sep['mean_salary']) * 100
    arrow2 = '↑' if delta2 > 0 else '↓' if delta2 < 0 else '→'
    color2 = 'green' if delta2 > 0 else 'red' if delta2 < 0 else 'gray'
    
    mid_x2 = (bars2[0].get_x() + bars2[0].get_width() + bars2[1].get_x()) / 2
    max_y2 = max(values2) + 15000
    
    ax2.text(mid_x2, max_y2, f'{arrow2} {delta2:+.1f}%', 
            ha='center', va='bottom', fontweight='bold', fontsize=12, color=color2)
    
    ax2.grid(True, alpha=0.3)
    
    # График 3: Топ-5 компаний за 5 октября
    if len(df_5oct) > 0:
        df_5oct_clean = categorize_roles(df_5oct)
        top_companies = df_5oct_clean['Компания'].value_counts().head(5)
        
        bars3 = ax3.barh(range(len(top_companies)), top_companies.values, color='lightcoral')
        ax3.set_yticks(range(len(top_companies)))
        ax3.set_yticklabels(top_companies.index)
        ax3.set_xlabel('Количество вакансий')
        ax3.set_title('🏢 Топ-5 компаний (5 октября)', fontsize=14, fontweight='bold')
        
        for i, bar in enumerate(bars3):
            width = bar.get_width()
            ax3.text(width + 0.1, bar.get_y() + bar.get_height()/2,
                    f'{int(width)}', ha='left', va='center', fontweight='bold')
    
    ax3.grid(True, alpha=0.3)
    
    # График 4: Распределение зарплат за 5 октября
    if len(df_5oct) > 0:
        df_5oct_clean = categorize_roles(df_5oct)
        df_with_salary = df_5oct_clean.dropna(subset=['salary_avg'])
        
        ax4.hist(df_with_salary['salary_avg'], bins=15, color='skyblue', edgecolor='navy', alpha=0.7)
        ax4.set_title('💰 Распределение зарплат (5 октября)', fontsize=14, fontweight='bold')
        ax4.set_xlabel('Зарплата, ₽')
        ax4.set_ylabel('Количество вакансий')
        
        # Статистики
        mean_salary = np.mean(df_with_salary['salary_avg'])
        median_salary = np.median(df_with_salary['salary_avg'])
        ax4.axvline(mean_salary, color='red', linestyle='--', linewidth=2, label=f'Средняя: {mean_salary:,.0f} ₽')
        ax4.axvline(median_salary, color='orange', linestyle='--', linewidth=2, label=f'Медиана: {median_salary:,.0f} ₽')
        ax4.legend()
    
    ax4.grid(True, alpha=0.3)
    
    plt.suptitle('📈 АВТОМАТИЧЕСКИЙ ОТЧЁТ ПО РЫНКУ ТРУДА ВЛАДИВОСТОКА\n26 сентября vs 5 октября 2025', 
                 fontsize=16, fontweight='bold')
    plt.tight_layout()
    plt.savefig(report_dir / 'summary_chart.png', dpi=300, bbox_inches='tight')
    plt.close()
    print("  ✅ Создан: summary_chart.png")

def load_data_by_date():
    """Загружает данные по датам"""
    data_26sep = []
    data_5oct = []
    
    # Данные за 26 сентября
    sep_dir = Path("data/2025-09-26")
    if sep_dir.exists():
        sep_files = list(sep_dir.glob("*.csv"))
        for csv_file in sep_files:
            try:
                df = pd.read_csv(csv_file, sep=';', encoding='utf-8-sig')
                data_26sep.append(df)
            except Exception as e:
                print(f"  ❌ Ошибка: {e}")
    
    # Данные за 5 октября
    oct_dir = Path("data/2025-10-05")
    if oct_dir.exists():
        oct_files = list(oct_dir.glob("*.csv"))
        for csv_file in oct_files:
            try:
                df = pd.read_csv(csv_file, sep=';', encoding='utf-8-sig')
                data_5oct.append(df)
            except Exception as e:
                print(f"  ❌ Ошибка: {e}")
    
    df_26sep = pd.concat(data_26sep, ignore_index=True) if data_26sep else pd.DataFrame()
    df_5oct = pd.concat(data_5oct, ignore_index=True) if data_5oct else pd.DataFrame()
    
    return df_26sep, df_5oct

def clean_salary_data(df, date_name):
    """Очищает данные о зарплатах"""
    if len(df) == 0:
        return df
    
    def parse_salary(salary_text):
        if pd.isna(salary_text) or salary_text == "не указано":
            return None, None
        
        salary_str = str(salary_text).replace(" ", "").replace(",", "").lower()
        
        range_match = re.search(r'(\d+)–(\d+)', salary_str)
        if range_match:
            from_salary = int(range_match.group(1))
            to_salary = int(range_match.group(2))
            return from_salary, to_salary
        
        from_match = re.search(r'от(\d+)', salary_str)
        if from_match:
            from_salary = int(from_match.group(1))
            return from_salary, None
        
        to_match = re.search(r'до(\d+)', salary_str)
        if to_match:
            to_salary = int(to_match.group(1))
            return None, to_salary
        
        return None, None
    
    salary_data = df['Зарплата'].apply(parse_salary)
    df['salary_from'] = salary_data.apply(lambda x: x[0] if x[0] is not None else np.nan)
    df['salary_to'] = salary_data.apply(lambda x: x[1] if x[1] is not None else np.nan)
    
    df['salary_avg'] = np.where(
        df['salary_to'].notna() & df['salary_from'].notna(),
        (df['salary_from'] + df['salary_to']) / 2,
        np.where(
            df['salary_from'].notna(),
            df['salary_from'],
            df['salary_to']
        )
    )
    
    df = df.dropna(subset=['salary_avg'])
    return df

def categorize_roles(df):
    """Определяет категории ролей"""
    if len(df) == 0:
        return df
    
    def get_role_category(title):
        if pd.isna(title):
            return "Неизвестно"
        
        title_lower = title.lower()
        
        sales_keywords = ["продаж", "sales", "менеджер по продаж", "руководитель продаж"]
        if any(keyword in title_lower for keyword in sales_keywords):
            return "Продажи"
        
        procurement_keywords = ["закуп", "закупк", "закупщик", "снабжен"]
        if any(keyword in title_lower for keyword in procurement_keywords):
            return "Закупки"
        
        project_keywords = ["проект", "project", "менеджер проект", "руководитель проект"]
        if any(keyword in title_lower for keyword in project_keywords):
            return "Проекты"
        
        management_keywords = ["менеджер", "руководитель", "директор"]
        if any(keyword in title_lower for keyword in management_keywords):
            return "Менеджмент"
        
        return "Другое"
    
    df['role_category'] = df['Название вакансии'].apply(get_role_category)
    return df

def get_stats(df, date_name):
    """Получает статистики для даты"""
    if len(df) == 0:
        return {
            'date': date_name,
            'total_vacancies': 0,
            'with_salary': 0,
            'unique_companies': 0,
            'mean_salary': 0,
            'median_salary': 0
        }
    
    salaries = df['salary_avg'].dropna()
    return {
        'date': date_name,
        'total_vacancies': len(df),
        'with_salary': len(salaries),
        'unique_companies': df['Компания'].nunique(),
        'mean_salary': np.mean(salaries),
        'median_salary': np.median(salaries)
    }

def create_excel_with_charts(report_dir):
    """Создаёт Excel файл с встроенными графиками"""
    print("📋 Создаём Excel файл с графиками...")
    
    # Создаём рабочую книгу
    wb = openpyxl.Workbook()
    
    # Удаляем дефолтный лист
    wb.remove(wb.active)
    
    # Создаём лист с данными
    ws_data = wb.create_sheet("📊 Данные и статистика")
    
    # Загружаем данные
    df_26sep, df_5oct = load_data_by_date()
    df_26sep = clean_salary_data(df_26sep, "26 сентября")
    df_5oct = clean_salary_data(df_5oct, "5 октября")
    
    stats_26sep = get_stats(df_26sep, "26 сентября")
    stats_5oct = get_stats(df_5oct, "5 октября")
    
    # Заголовок
    ws_data['A1'] = "📈 АВТОМАТИЧЕСКИЙ ОТЧЁТ ПО РЫНКУ ТРУДА ВЛАДИВОСТОКА"
    ws_data['A1'].font = Font(size=16, bold=True)
    ws_data.merge_cells('A1:D1')
    
    # Дата создания
    ws_data['A2'] = f"Дата создания: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws_data['A2'].font = Font(size=12, italic=True)
    
    # Сводная таблица
    ws_data['A4'] = "📊 СРАВНИТЕЛЬНАЯ СТАТИСТИКА"
    ws_data['A4'].font = Font(size=14, bold=True)
    
    # Заголовки таблицы
    headers = ['Показатель', '26 сентября', '5 октября', 'Изменение']
    for i, header in enumerate(headers, 1):
        cell = ws_data.cell(row=6, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Данные
    data_rows = [
        ['Количество вакансий с зарплатой', stats_26sep['with_salary'], stats_5oct['with_salary']],
        ['Уникальных компаний', stats_26sep['unique_companies'], stats_5oct['unique_companies']],
        ['Средняя зарплата', f"{stats_26sep['mean_salary']:,.0f} ₽", f"{stats_5oct['mean_salary']:,.0f} ₽"],
        ['Медианная зарплата', f"{stats_26sep['median_salary']:,.0f} ₽", f"{stats_5oct['median_salary']:,.0f} ₽"]
    ]
    
    for i, row_data in enumerate(data_rows, 7):
        for j, value in enumerate(row_data, 1):
            ws_data.cell(row=i, column=j, value=value)
        
        # Вычисляем изменение для числовых показателей
        if i <= 8:  # Для первых двух строк (количественные показатели)
            old_val = row_data[1]
            new_val = row_data[2]
            if old_val > 0:
                change_pct = ((new_val - old_val) / old_val) * 100
                arrow = '↑' if change_pct > 0 else '↓' if change_pct < 0 else '→'
                ws_data.cell(row=i, column=4, value=f"{arrow} {change_pct:+.1f}%")
            else:
                ws_data.cell(row=i, column=4, value="—")
    
    # Настройка ширины столбцов
    ws_data.column_dimensions['A'].width = 30
    ws_data.column_dimensions['B'].width = 20
    ws_data.column_dimensions['C'].width = 20
    ws_data.column_dimensions['D'].width = 15
    
    # Создаём лист с графиками
    ws_charts = wb.create_sheet("📊 Графики")
    
    # Вставляем график
    chart_path = report_dir / 'summary_chart.png'
    if chart_path.exists():
        img = ExcelImage(chart_path)
        img.width = 800
        img.height = 600
        ws_charts.add_image(img, 'A1')
        
        # Заголовок листа
        ws_charts['A12'] = "📈 Сводная аналитика рынка труда Владивостока"
        ws_charts['A12'].font = Font(size=14, bold=True)
        
        # Описание
        ws_charts['A13'] = "• График 1: Динамика количества вакансий с зарплатой"
        ws_charts['A14'] = "• График 2: Динамика средних зарплат"
        ws_charts['A15'] = "• График 3: Топ-5 компаний по количеству вакансий (5 октября)"
        ws_charts['A16'] = "• График 4: Распределение зарплат (5 октября)"
    
    # Сохраняем файл
    excel_file = report_dir / 'automated_report.xlsx'
    wb.save(excel_file)
    print(f"  ✅ Создан: {excel_file}")

def main():
    """Основная функция"""
    print("🤖 СОЗДАНИЕ АВТОМАТИЧЕСКОГО ОТЧЁТА")
    print("=" * 50)
    
    # Создаём папку для отчёта
    report_dir = Path("report_automated")
    report_dir.mkdir(exist_ok=True)
    print(f"📁 Папка для автоматического отчёта: {report_dir.absolute()}")
    
    try:
        # 1. Создаём сводный график
        create_summary_chart(report_dir)
        
        # 2. Создаём Excel с графиками
        create_excel_with_charts(report_dir)
        
        print("\n🎉 АВТОМАТИЧЕСКИЙ ОТЧЁТ СОЗДАН!")
        print(f"📁 Все файлы сохранены в папке: {report_dir.absolute()}")
        print("\n📊 Созданные файлы:")
        print("  • summary_chart.png - Сводный график")
        print("  • automated_report.xlsx - Excel отчёт с графиками")
        
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        return False
    
    return True

if __name__ == "__main__":
    main()

